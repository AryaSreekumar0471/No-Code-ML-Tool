from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, session, redirect, url_for, Response
from sklearn.model_selection import train_test_split
from imblearn.under_sampling import RandomUnderSampler
from imblearn.over_sampling import SMOTE
from imblearn.pipeline import Pipeline
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import cross_val_score, StratifiedKFold
from sklearn.metrics import precision_score, recall_score, f1_score, make_scorer
from dtreeviz import dtreeviz
from PIL import Image
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import confusion_matrix
from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.calibration import calibration_curve
from sklearn.metrics import roc_curve, auc, precision_recall_curve
from openpyxl.styles import Alignment

import pickle
import plotly.express as px
import plotly.graph_objs as go
from sklearn.model_selection import cross_val_predict
from sklearn.metrics import confusion_matrix
from sklearn import tree
from sklearn.tree import DecisionTreeClassifier
import plotly.figure_factory as ff
import plotly

import base64
import os
import io
import json
import matplotlib
matplotlib.use("Agg") 
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler, RobustScaler, LabelEncoder, MinMaxScaler
from sklearn.impute import KNNImputer

app = Flask(__name__)
app.secret_key = "replace_this_with_a_secure_random_key"
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
uploaded_file_path = None


# ===== Helpers for experiment page =====
def _normalize_names(cols):
    return (pd.Index(cols)
              .str.strip()
              .str.lower()
              .str.replace(' ', '_'))

def _build_schema_from_df(df, cols):
    schema = []
    for c in cols:
        if c not in df.columns:
            continue
        s = df[c]
        if pd.api.types.is_numeric_dtype(s):
            s_num = pd.to_numeric(s, errors='coerce')
            minv = float(np.nanmin(s_num))
            maxv = float(np.nanmax(s_num))

            # integer-like if dtype is int OR all non-nan values are whole numbers
            integer_like = (
                pd.api.types.is_integer_dtype(s) or
                (s_num.dropna() % 1 == 0).all()
            )
            step = 1 if integer_like else 'any'

            # sensible default: median; round to int for integer-like
            default = float(np.nanmedian(s_num))
            if integer_like:
                default = int(round(default))

            schema.append({
                "name": c,
                "label": c.replace('_', ' ').title(),
                "type": "number",
                "min": minv,
                "max": maxv,
                "step": step,
                "default": default,
            })
        else:
            schema.append({
                "name": c,
                "label": c.replace('_', ' ').title(),
                "type": "categorical",
                "options": sorted([str(x) for x in s.dropna().unique().tolist()])
            })
    return schema

def _payload_to_model_row(payload):
    steps = (session.get('preprocess') or {}).get('steps', [])
    encode_method = session.get('encode_method_used', 'onehot')
    onehot_categories = session.get('onehot_categories') or {}
    label_mapping = session.get('label_mapping') or {}
    scale_params = session.get('scale_params') or {}
    X_cols = session.get('selected_features', [])

    # 1) one-row DataFrame with ORIGINAL column names
    row = pd.DataFrame([payload])

    # apply the same "Clean column names" if used
    if 'clean' in steps:
        row.columns = _normalize_names(row.columns)

    # coerce dtypes
    for c in row.columns:
        if c in label_mapping or c in onehot_categories:
            row[c] = row[c].astype(str)
        else:
            row[c] = pd.to_numeric(row[c], errors='coerce')

    # 2) encode
    if 'encode' in steps:
        if encode_method == 'label':
            for col, mapping in label_mapping.items():
                if col in row.columns:
                    val = str(row.at[0, col])
                    row.at[0, col] = mapping.get(val, np.nan)
            enc_df = row
        else:
            enc_df = row.copy()
            for col, cats in onehot_categories.items():
                for cat in cats:
                    enc_df[f"{col}_{cat}"] = 1 if str(row.get(col, "")) == str(cat) else 0
                if col in enc_df.columns:
                    enc_df.drop(columns=[col], inplace=True)
    else:
        enc_df = row

    # 3) scale
    if scale_params.get('nums'):
        if scale_params['method'] == 'normalize':
            mins = dict(zip(scale_params['nums'], scale_params['min']))
            maxs = dict(zip(scale_params['nums'], scale_params['max']))
            for c in scale_params['nums']:
                if c in enc_df.columns:
                    a = pd.to_numeric(enc_df[c], errors='coerce')
                    den = (maxs[c] - mins[c]) or 1.0
                    enc_df[c] = (a - mins[c]) / den
        else:
            means = dict(zip(scale_params['nums'], scale_params['mean']))
            scales = dict(zip(scale_params['nums'], scale_params['scale']))
            for c in scale_params['nums']:
                if c in enc_df.columns:
                    a = pd.to_numeric(enc_df[c], errors='coerce')
                    den = scales[c] or 1.0
                    enc_df[c] = (a - means[c]) / den

    # 4) align to model feature order (missing cols -> 0)
    final_row = enc_df.reindex(columns=X_cols, fill_value=0)
    return final_row


@app.route('/')
def home():
    return render_template('index.html')

@app.route('/upload', methods=['GET'])
def upload_page():
    preview = None
    if request.args.get('preview') == 'true' and uploaded_file_path:
        df = pd.read_csv(uploaded_file_path)
        preview = df.head().to_html(
            classes='table table-striped table-bordered mt-3',
            index=False
        )
    return render_template('upload.html', preview=preview)

@app.route('/upload_file', methods=['POST'])
def upload_file():
    global uploaded_file_path
    file = request.files.get('dataset')
    if file and file.filename.lower().endswith('.csv'):
        save_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(save_path)
        uploaded_file_path = save_path 
        #remember dataset name + raw shape
        # ✅ Clear all old session info tied to the previous dataset
        for k in [
            'preprocess_summary', 'preprocess', 'cleaned_file',
            'balance_method', 'class_counts_before', 'class_counts_after',
            'selected_features', 'selected_target',
            'train_file', 'test_file', 'train_file_orig',
            'cv_folds',
            'encode_method_used', 'onehot_categories', 'label_mapping',
            'encoded_to_source', 'scale_params',
            'split_ratio', 'train_size', 'test_size'
        ]:
            session.pop(k, None)
      

        session['dataset_name'] = file.filename
        # Clear old preprocessing/session info so skipping preprocessing doesn't show old steps
        session.pop('preprocess_summary', None)
        session.pop('preprocess', None)
        session.pop('cleaned_file', None)
        session.pop('balance_method', None)
        session.pop('class_counts_before', None)
        session.pop('class_counts_after', None)
        try:
            df_tmp = pd.read_csv(save_path, nrows=500000)  # prevents accidental huge load
            session['n_rows_total'] = int(df_tmp.shape[0])
            session['n_cols_total'] = int(df_tmp.shape[1])
        except Exception:
            session['n_rows_total'] = None
            session['n_cols_total'] = None

        return jsonify(success=True), 200
    return jsonify(success=False), 400

@app.route('/delete_file', methods=['POST'])
def delete_file():
    global uploaded_file_path
    filename = request.json.get('filename', '')
    path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(path):
        os.remove(path)
    uploaded_file_path = None
    # Clear all old session info tied to the previous dataset
        
    for k in [
            'preprocess_summary', 'preprocess', 'cleaned_file',
            'balance_method', 'class_counts_before', 'class_counts_after',
            'selected_features', 'selected_target',
            'train_file', 'test_file', 'train_file_orig',
            'cv_folds',
            'encode_method_used', 'onehot_categories', 'label_mapping',
            'encoded_to_source', 'scale_params',
            'split_ratio', 'train_size', 'test_size'
        ]:
            session.pop(k, None)
        #  End clear
    return jsonify(deleted=True), 200

@app.route('/preprocess_page')
def preprocess_page():
    return render_template('preprocess.html')

@app.route('/missing_data_data')
def missing_data_data():
    df = pd.read_csv(uploaded_file_path)
    counts = df.isna().sum().to_dict()
    return jsonify(data=counts)

@app.route('/categorical_count_data')
def categorical_count_data():
    df = pd.read_csv(uploaded_file_path)
    result = {}
    for col in df.select_dtypes(include=['object', 'category']).columns:
        counts = df[col].fillna('Missing').value_counts().to_dict()
        result[col] = counts
    return jsonify(data=result)

@app.route('/boxplot')
def boxplot():
    df = pd.read_csv(uploaded_file_path)
    nums = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
    if not nums:
        return ('', 204)
    fig, ax = plt.subplots(figsize=(len(nums)*0.8, 4))
    df[nums].boxplot(ax=ax, vert=False)
    ax.set_title("Box Plots of Numeric Features")
    ax.tick_params(axis='y', labelsize=8)
    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return send_file(buf, mimetype='image/png')

@app.route('/run_preprocessing', methods=['POST'])
def run_preprocessing():
    global uploaded_file_path
    body = request.get_json()
    steps = body.get('steps', [])
    encode_method = body.get('encode_method')
    outlier_method = body.get('outlier_method')

    df = pd.read_csv(uploaded_file_path)
    summary = []

    if 'clean' in steps:
        df.columns = (df.columns
                      .str.strip()
                      .str.lower()
                      .str.replace(' ', '_'))
        summary.append("Cleaned column names")

    if 'missing' in steps:
        before = df.shape[0]
        df = df.dropna()
        after = df.shape[0]
        summary.append(f"Dropped missing: {before} → {after} rows")

    if 'duplicates' in steps:
        before = df.shape[0]
        df = df.drop_duplicates()
        after = df.shape[0]
        summary.append(f"Dropped duplicates: {before} → {after} rows")
    
        # keep a copy before encoding so we know original categories
    pre_encode_df = df.copy()
    session['original_columns_for_experiment'] = pre_encode_df.columns.tolist()

    # ENCODE
    # -------------------------
    onehot_categories = {}
    label_mapping = {}
    encoded_to_source = {}

    if 'encode' in steps:
        if encode_method == 'label':
            for col in pre_encode_df.select_dtypes(include=['object', 'category']).columns:
                le = LabelEncoder()
                le.fit(pre_encode_df[col].astype(str))
                df[col] = le.transform(df[col].astype(str))
                label_mapping[col] = {cls: int(i) for i, cls in enumerate(le.classes_.astype(str))}
                encoded_to_source[col] = col
            summary.append("Label-encoded categorical features")
        else:
            cat_cols = pre_encode_df.select_dtypes(include=['object', 'category']).columns
            for col in cat_cols:
                cats = sorted(pre_encode_df[col].astype(str).dropna().unique().tolist())
                onehot_categories[col] = cats
            df = pd.get_dummies(df)
            for src, cats in onehot_categories.items():
                for cat in cats:
                    encoded_to_source[f"{src}_{cat}"] = src
            summary.append("One-Hot encoded categorical features")

        # persist encoding meta
        session['encode_method_used'] = encode_method or 'onehot'
        session['onehot_categories'] = onehot_categories
        session['label_mapping'] = label_mapping
        session['encoded_to_source'] = encoded_to_source

        
# SCALE
# -------------------------
    if 'scale' in steps:
        target = session.get('selected_target')  # may be None at this stage

        # all numeric
        all_nums = df.select_dtypes(include=['int64', 'float64']).columns.tolist()

        # if target already known, exclude it directly
        explicit_exclude = set()
        if target and target in all_nums:
            explicit_exclude.add(target)

        # If target is NOT known yet, protect columns that look like labels
        # Heuristic: integer dtype + few unique values (≤10 or ≤5% of rows)
        n_rows = len(df)
        uniq_cap = max(2, min(10, int(0.05 * n_rows)))  # cap based on dataset size
        label_like = []
        for c in all_nums:
            if pd.api.types.is_integer_dtype(df[c]):
                nunq = df[c].nunique(dropna=True)
                if nunq <= uniq_cap:
                    label_like.append(c)

        # Build final list to scale: numeric minus explicit target minus label-like
        protected = explicit_exclude.union(label_like)
        nums = [c for c in all_nums if c not in protected]

        if not nums:
            summary.append("No numeric features to scale (target/label-like columns protected).")
        else:
            method = body.get('scale_method', 'standardize')
            if method == 'normalize':
                scaler = MinMaxScaler().fit(df[nums])
                df[nums] = scaler.transform(df[nums])
                session['scale_params'] = {
                    'method': 'normalize',
                    'nums': nums,
                    'min': scaler.data_min_.tolist(),
                    'max': scaler.data_max_.tolist()
                }
                summary.append(f"Normalized numeric features (min–max): {', '.join(nums)}")
            else:
                scaler = StandardScaler().fit(df[nums])
                df[nums] = scaler.transform(df[nums])
                session['scale_params'] = {
                    'method': 'standardize',
                    'nums': nums,
                    'mean': scaler.mean_.tolist(),
                    'scale': scaler.scale_.tolist()
            }
                summary.append(f"Standardized numeric features (z-score): {', '.join(nums)}")

        # Tell the user what we deliberately did NOT scale (helps transparency)
        if protected:
            summary.append(
                "Skipped scaling on possible label columns: " + ", ".join(sorted(protected))
        )
    

    if 'outliers' in steps:
        nums = [col for col in df.select_dtypes(include=['int64', 'float64']).columns if df[col].nunique() > 2]
        Q1 = df[nums].quantile(0.25)
        Q3 = df[nums].quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR
        mask = (df[nums] < lower) | (df[nums] > upper)

        
        if outlier_method == 'remove':
            before = df.shape[0]
            df = df[~mask.any(axis=1)]
            summary.append(f"Removed outliers: {before} → {df.shape[0]} rows")
        elif outlier_method == 'cap':
             df[nums] = df[nums].clip(lower=lower, upper=upper, axis=1)
             summary.append("Capped outliers to 1.5×IQR bounds")

        else: 
            # Default to robust if not specified or invalid
             scaler = RobustScaler()
             scaled = scaler.fit_transform(df[nums])
             scaled_df = pd.DataFrame(scaled, columns=nums)

             # ( scaled_df[mask] = np.nan here)
             imputer = KNNImputer(n_neighbors=5)
             imputed = imputer.fit_transform(scaled_df)

             df[nums] = scaler.inverse_transform(imputed)
             summary.append("Applied RobustScaler + KNN imputation on numeric features")

        

    cleaned_path = uploaded_file_path.replace('.csv', '_cleaned.csv')
    df.to_csv(cleaned_path, index=False)
    preview_html = df.head().to_html(classes="table table-striped", index=False)

    #persist selections & summary for the summary card
    session['preprocess'] = {
       "steps": steps,                      # e.g. ["clean","missing","encode","scale","outliers"]
       "encode_method": encode_method,      # "label" | "onehot"
       "scale_method": body.get('scale_method', 'standardize'),  # "normalize" | "standardize"
       "outlier_method": outlier_method,    # "remove" | "cap" | "robust" | None
    }
    session['preprocess_summary'] = summary   # your human-readable bullet list
    session['cleaned_file'] = os.path.basename(cleaned_path)

    return jsonify(status='done', summary=summary, preview=preview_html, cleaned_file=os.path.basename(cleaned_path))

@app.route('/download_cleaned/<filename>')
def download_cleaned(filename):
    folder = os.path.dirname(uploaded_file_path)
    return send_from_directory(directory=folder, path=filename, as_attachment=True, mimetype='text/csv')

@app.route('/select_features', methods=['GET', 'POST'])
def select_features():
    # Prefer the cleaned file if it exists; otherwise use the original upload
    cleaned_path = uploaded_file_path.replace('.csv', '_cleaned.csv') if uploaded_file_path else None
    path = cleaned_path if (cleaned_path and os.path.exists(cleaned_path)) else uploaded_file_path

    if request.method == 'POST':
        data = request.get_json()
        feats = data.get('features', [])
        targ  = data.get('target', None)
        feats = [f for f in feats if f != targ]
        if not feats or not targ:
            return jsonify(success=False, message="Select at least one feature and exactly one target"), 400
        session['selected_features'] = feats
        session['selected_target']   = targ
        return jsonify(success=True), 200

    df = pd.read_csv(path)
    cols = df.columns.tolist()
    return render_template('select_features.html', columns=cols)

@app.route('/feature_distribution_data/<path:feature>')
def feature_distribution_data(feature):
    # Prefer the cleaned file if it exists; otherwise use the original upload
    cleaned_path = uploaded_file_path.replace('.csv', '_cleaned.csv') if uploaded_file_path else None
    path = cleaned_path if (cleaned_path and os.path.exists(cleaned_path)) else uploaded_file_path

    df = pd.read_csv(path)
    series = df[feature]
    target = session.get('selected_target')

    if feature == target:
        counts = series.value_counts().to_dict()
        return jsonify(type='categorical', counts=counts)

    if pd.api.types.is_numeric_dtype(series):
        raw_counts, raw_bins = np.histogram(series.dropna(), bins=10)
        return jsonify(type='numeric',
                       bins=[float(x) for x in raw_bins],
                       counts=[int(x) for x in raw_counts])
    else:
        counts = series.fillna('Missing').value_counts().to_dict()
        return jsonify(type='categorical', counts=counts)

@app.route('/data_overview', methods=['GET', 'POST'])
def data_overview():
    global uploaded_file_path
    # Prefer the cleaned file if it exists; otherwise use the original upload
    cleaned_path = uploaded_file_path.replace('.csv', '_cleaned.csv') if uploaded_file_path else None
    path = cleaned_path if (cleaned_path and os.path.exists(cleaned_path)) else uploaded_file_path

    df = pd.read_csv(path)
    total_rows = len(df)

    if request.method == 'POST':
        if request.form.get('action') == 'reselect':
            return redirect(url_for('select_features'))

        if 'ratio' in request.form:
            ratio = int(request.form['ratio'])
            train_size = ratio / 100.0
            target_col = session.get('selected_target')

            # Stratified split
            train_df, test_df = train_test_split(
                df,
                train_size=train_size,
                random_state=42,
                stratify=df[target_col]
            )

            # Save next to whichever file we're actually using (cleaned or original)
            base = os.path.splitext(path)[0]
            train_path = f"{base}_train.csv"
            test_path  = f"{base}_test.csv"
            train_df.to_csv(train_path, index=False)
            test_df.to_csv(test_path, index=False)

            session['train_file'] = os.path.basename(train_path)
            session['test_file']  = os.path.basename(test_path)

            # Store summary for later pages
            session['split_ratio'] = {"train": train_size, "test": 1 - train_size}
            session['train_size']  = int(train_df.shape[0])
            session['test_size']   = int(test_df.shape[0])

            return redirect(url_for('class_balance'))

    features    = session.get('selected_features', [])
    target_col  = session.get('selected_target', '')
    default_ratio = 70

    return render_template(
        'data_overview.html',
        features=features,
        target=target_col,
        default_ratio=default_ratio,
        total_rows=total_rows
    )

@app.route('/class-balance', methods=['GET']) 
def class_balance():
    # Remember the original (unbalanced) file once
    if 'train_file_orig' not in session and session.get('train_file'):
        session['train_file_orig'] = session['train_file']

    # On refresh/open: revert to original unbalanced data and clear any "after" info
    if session.get('train_file_orig'):
        orig = session['train_file_orig']
        if orig.endswith('_balanced.csv'):
            orig = orig[:-len('_balanced.csv')] + '.csv'
        # DO NOT strip _cleaned_train.csv
        session['train_file'] = orig
        session['train_file_orig'] = orig

    session.pop('balance_method', None)
    session.pop('class_counts_after', None)

    # Load counts from the original file
    train_filename = session.get('train_file')
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)
    df = pd.read_csv(train_path)
    target = session.get('selected_target')

    # --- NEW: resolve target against actual columns (case/space/BOM) ---
    def _norm(s):
        return str(s).lstrip('\ufeff').strip().lower() if s is not None else s
    col_map = {_norm(c): c for c in df.columns}
    t_key = _norm(target)
    if t_key in col_map:
        target = col_map[t_key]
        session['selected_target'] = target
    else:
        # optional single-column autodetect (integer with few uniques)
        n_rows = len(df)
        uniq_cap = max(2, min(10, int(0.05 * n_rows)))
        candidates = [c for c in df.columns
                      if pd.api.types.is_integer_dtype(df[c]) and df[c].nunique(dropna=True) <= uniq_cap]
        if len(candidates) == 1:
            target = candidates[0]
            session['selected_target'] = target
        else:
            raise KeyError(f"Target '{session.get('selected_target')}' not found in this dataset.")
    # --- END NEW ---

    counts = df[target].value_counts().to_dict()
    session['class_counts_before'] = {str(k): int(v) for k, v in counts.items()}

    return render_template('class_balance.html', counts=counts)


@app.route('/class-balance-data', methods=['POST'])
def class_balance_data():
    data = request.get_json()
    method = data.get('method')
    percent = int(data.get('percent', 50))

    train_filename = session.get('train_file')

    # Rebase only if we were on a balanced file; keep _cleaned_train if present
    if train_filename.endswith('_balanced.csv'):
        train_filename = train_filename[:-len('_balanced.csv')] + '.csv'
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)


    df = pd.read_csv(train_path)
    target = session.get('selected_target')

    # --- NEW: resolve target against actual columns (case/space/BOM) ---
    def _norm(s):
        return str(s).lstrip('\ufeff').strip().lower() if s is not None else s
    col_map = {_norm(c): c for c in df.columns}
    t_key = _norm(target)
    if t_key in col_map:
        target = col_map[t_key]
        session['selected_target'] = target
    else:
        n_rows = len(df)
        uniq_cap = max(2, min(10, int(0.05 * n_rows)))
        candidates = [c for c in df.columns
                      if pd.api.types.is_integer_dtype(df[c]) and df[c].nunique(dropna=True) <= uniq_cap]
        if len(candidates) == 1:
            target = candidates[0]
            session['selected_target'] = target
        else:
            return jsonify(error=f"Target '{session.get('selected_target')}' not found in this dataset."), 400
    # --- END NEW ---

    X = df.drop(columns=[target])
    y = df[target]

    try:
        if method == 'under':
            rus = RandomUnderSampler(sampling_strategy='auto', random_state=42)
            X_res, y_res = rus.fit_resample(X, y)

        elif method == 'over':
            smote = SMOTE(sampling_strategy='auto', random_state=42)
            X_res, y_res = smote.fit_resample(X, y)

        else:
            # BOTH: apply to ALL classes (multiclass-safe)
            counts = y.value_counts()

            # target count after balancing (e.g., 50% of current max when percent=50)
            target_n = int(counts.max() * (percent / 100.0))
            target_n = max(target_n, 2)  # avoid degenerate sizes

            # 1) UNDER-SAMPLE: cap every class at target_n
            under_strategy = {cls: min(int(cnt), target_n) for cls, cnt in counts.items()}
            rus = RandomUnderSampler(sampling_strategy=under_strategy, random_state=42)
            X_under, y_under = rus.fit_resample(X, y)

            # 2) OVER-SAMPLE: lift any class below target_n up to target_n
            counts_under = pd.Series(y_under).value_counts()
            over_strategy = {cls: int(target_n) for cls, cnt in counts_under.items() if int(cnt) < target_n}

            if over_strategy:
                # choose a safe k for smallest class after undersampling
                min_class_size = int(counts_under.min())
                k_safe = max(1, min(5, min_class_size - 1))
                smote = SMOTE(sampling_strategy=over_strategy, random_state=42, k_neighbors=k_safe)
                X_res, y_res = smote.fit_resample(X_under, y_under)
            else:
                # nothing to oversample; already balanced at target_n
                X_res, y_res = X_under, y_under

        new_counts = pd.Series(y_res).value_counts().to_dict()

        # Persist choice and counts
        session['balance_method'] = method
        session['class_counts_after'] = {str(k): int(v) for k, v in new_counts.items()}

        # Save balanced training data to a new file and point session to it
        balanced_df = pd.concat([pd.DataFrame(X_res), pd.Series(y_res, name=target)], axis=1)
        base = os.path.splitext(train_path)[0]
        balanced_path = f"{base}_balanced.csv"
        balanced_df.to_csv(balanced_path, index=False)
        session['train_file'] = os.path.basename(balanced_path)

        return jsonify(counts=new_counts)

    except Exception as e:
        return jsonify(error=str(e)), 400

@app.route('/set_cv_folds', methods=['POST'])
def set_cv_folds():
    data = request.get_json() or {}
    try:
        folds = int(data.get('folds', 5))
    except:
        folds = 5
    session['cv_folds'] = folds
    return jsonify(success=True)

@app.route('/model_selection')
def model_selection():
    return render_template('model_selection.html')

@app.route('/train_decision_tree')
def train_decision_tree():
    return render_template('train_decision_tree.html')

@app.route('/api/train_decision_tree', methods=['POST'])
def api_train_decision_tree():
    features = session.get('selected_features', [])
    target = session.get('selected_target', None)
    cv_folds = int(session.get('cv_folds', 5))
    train_filename = session.get('train_file')
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)

    df = pd.read_csv(train_path)
    X = df[features]
    y = df[target]

    clf = DecisionTreeClassifier(random_state=42)
    cv = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=42)

    accuracy_scores = cross_val_score(clf, X, y, cv=cv, scoring='accuracy')
    precision_scores = cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(precision_score, average='macro', zero_division=0))
    recall_scores = cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(recall_score, average='macro', zero_division=0))
    f1_scores = cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(f1_score, average='macro', zero_division=0))

    result = {
        "accuracy": f"{np.mean(accuracy_scores):.2f}",
        "accuracy_std": f"{np.std(accuracy_scores):.2f}",
        "precision": f"{np.mean(precision_scores):.2f}",
        "recall": f"{np.mean(recall_scores):.2f}",
        "f1": f"{np.mean(f1_scores):.2f}"
    }
    return jsonify(result)

@app.route('/download_results', methods=['POST'])
def download_results():
    data = request.get_json(silent=True) or {}
    metrics = data.get('metrics', {})

    # ---- Gather from session
    model_name   = session.get('model_name', 'Decision Tree')
    dataset_name = session.get('dataset_name') or session.get('cleaned_file') or session.get('train_file', '—')
    features     = session.get('selected_features', [])
    target       = session.get('selected_target', '—')
    cv_folds     = session.get('cv_folds', 5)
    sr           = session.get('split_ratio') or {}
    train_size   = session.get('train_size', '—')
    test_size    = session.get('test_size', '—')
    preproc_list = session.get('preprocess_summary', [])
    method_code  = session.get('balance_method')
    before       = session.get('class_counts_before', {})
    after        = session.get('class_counts_after', {})

    # Friendly class-balance method
    method_map = {'under': 'Undersampling', 'over': 'Oversampling', 'both': 'Both over and under sampling'}
    balance_method = method_map.get(method_code, 'None' if not method_code else method_code)

    # Split ratio string (e.g., 80 : 20)
    if sr and 'train' in sr and 'test' in sr:
        split_ratio_str = f"{int(round(sr['train']*100))} : {int(round(sr['test']*100))}"
    else:
        split_ratio_str = '—'

    # ---- Build the all-in-one “Report” rows
    rows = []
    add = rows.append

    # Section: Model Summary
    add(('Model Summary', 'Model', model_name))
    add(('Model Summary', 'Dataset', dataset_name))
    add(('Model Summary', 'Target', target))
    add(('Model Summary', 'Features', ', '.join(features)))
    add(('Model Summary', 'Cross-Validation Folds', cv_folds))

    # Section: Data Split
    add(('Data Split', 'Train/Test Split', split_ratio_str))
    add(('Data Split', 'Train set size', train_size))
    add(('Data Split', 'Test set size', test_size))

    # Section: Class Balancing
    add(('Class Balancing', 'Method', balance_method))
    if before:
        for k, v in sorted(before.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'Before - Class {k}', v))
    if after:
        for k, v in sorted(after.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'After - Class {k}', v))

    # Section: Preprocessing
    if preproc_list:
        for i, step in enumerate(preproc_list, 1):
            add(('Preprocessing', f'Step {i}', step))
    else:
        add(('Preprocessing', 'Steps', 'None'))

    # Section: Results (metrics from the Train action)
    add(('Results', 'Accuracy',      metrics.get('accuracy')))
    add(('Results', 'Accuracy Std',  metrics.get('accuracy_std')))
    add(('Results', 'Precision',     metrics.get('precision')))
    add(('Results', 'Recall',        metrics.get('recall')))
    add(('Results', 'F1-score',      metrics.get('f1')))

    df_report = pd.DataFrame(rows, columns=['Section', 'Field', 'Value'])

    # Optional: keep separate sheets too
    df_summary = df_report[df_report['Section'] == 'Model Summary'][['Field', 'Value']]
    df_metrics = df_report[df_report['Section'] == 'Results'][['Field', 'Value']].T
    df_metrics.columns = df_metrics.iloc[0]; df_metrics = df_metrics[1:].reset_index(drop=True)

    if preproc_list:
        df_preproc = pd.DataFrame({'Preprocessing Steps': preproc_list})
    else:
        df_preproc = pd.DataFrame({'Preprocessing Steps': ['None']})

    classes = sorted(set(list(before.keys()) + list(after.keys())))
    df_balance = pd.DataFrame({
        'Class': classes,
        'Before': [before.get(c, 0) for c in classes],
        'After':  [after.get(c, 0) for c in classes],
    })

    # ---- Write Excel (openpyxl)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # All-in-one first
        df_report.to_excel(writer, index=False, sheet_name='Report (All-in-one)')

        # Optional individual tabs (remove if you truly want only one sheet)
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        df_metrics.to_excel(writer, index=False, sheet_name='Metrics')
        df_preproc.to_excel(writer, index=False, sheet_name='Preprocessing')
        df_balance.to_excel(writer, index=False, sheet_name='Class Balance')
        pd.DataFrame({'Features': features}).to_excel(writer, index=False, sheet_name='Features')

        # Widen columns on the Report sheet
        ws = writer.sheets['Report (All-in-one)']
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 80

        # NEW: left-align everything on all sheets
        left = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # All-in-one sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = left

        # Other sheets
        for name in ['Summary', 'Metrics', 'Preprocessing', 'Class Balance', 'Features']:
            if name in writer.sheets:
                s = writer.sheets[name]
                for row in s.iter_rows(min_row=1, max_row=s.max_row, min_col=1, max_col=s.max_column):
                    for cell in row:
                        cell.alignment = left

    output.seek(0)
    filename = f"{model_name.replace(' ', '_').lower()}_results.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/plot/decision_tree')
def plot_decision_tree():
    train_filename = session['train_file']
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_filename))
    X = df[session['selected_features']]
    y = df[session['selected_target']]

    clf = DecisionTreeClassifier(random_state=42).fit(X, y)
    fig_mpl, ax = plt.subplots(figsize=(16, 6), dpi=100)
    tree.plot_tree(
        clf,
        max_depth=4,
        feature_names=X.columns,
        class_names=[str(c) for c in np.unique(y)],
        filled=True,
        rounded=True,
        fontsize=10,
        ax=ax
    )
    plt.tight_layout()

    buf = io.BytesIO()
    fig_mpl.savefig(buf, format='png', dpi=100)
    plt.close(fig_mpl)
    buf.seek(0)

    img = Image.open(buf)
    img_arr = np.array(img)

    fig = px.imshow(img_arr)
    fig.update_xaxes(visible=False).update_yaxes(visible=False)
    fig.update_layout(
        title="Decision Tree (Top 4 Levels)",
        margin=dict(l=10, r=10, t=40, b=10),
        autosize=True
    )

    js = fig.to_plotly_json()
    return jsonify(data=js["data"], layout=js["layout"])

@app.route('/plot/feature_importance')
def plot_feature_importance():
    train_filename = session['train_file']
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)
    df = pd.read_csv(train_path)

    X = df[session['selected_features']]
    y = df[session['selected_target']]

    clf = DecisionTreeClassifier(random_state=42)
    clf.fit(X, y)

    feature_names = X.columns.tolist()
    importances = clf.feature_importances_.tolist()

    idx = sorted(range(len(importances)), key=lambda i: importances[i], reverse=True)
    names_sorted = [feature_names[i] for i in idx]
    imps_sorted = [importances[i] for i in idx]

    fig = go.Figure(go.Bar(
        x=imps_sorted,
        y=names_sorted,
        orientation='h',
        marker_color='rgba(0,123,255,0.8)'
    ))

    fig.update_layout(
        title="Feature Importance",
        xaxis_title="Importance",
        yaxis_title="Feature",
        yaxis=dict(autorange="reversed"),
        margin=dict(l=200, r=50, t=50, b=50),
        height=500
    )

    payload = {
        "data": fig.to_plotly_json()["data"],
        "layout": fig.to_plotly_json()["layout"]
    }
    body = json.dumps(payload, cls=plotly.utils.PlotlyJSONEncoder)
    return Response(body, mimetype='application/json')

@app.route('/plot/confusion_matrix')
def plot_confusion_matrix():
    train_filename = session['train_file']
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)
    df = pd.read_csv(train_path)

    X = df[session['selected_features']]
    y = df[session['selected_target']]
    preds = cross_val_predict(
        DecisionTreeClassifier(random_state=42),
        X, y,
        cv=int(session['cv_folds'])
    )

    cm = confusion_matrix(y, preds)
    labels = [str(c) for c in np.unique(y)]
    z = cm.tolist()

    fig = go.Figure(go.Heatmap(
        z=z,
        x=labels,
        y=labels,
        colorscale="Blues",
        hoverongaps=False,
        showscale=True
    ))
    fig.update_layout(
        title="Confusion Matrix",
        xaxis=dict(title="Predicted", type="category"),
        yaxis=dict(title="Actual", type="category", autorange="reversed"),
        margin=dict(l=60, r=60, t=50, b=50)
    )

    payload = fig.to_plotly_json()
    body = json.dumps(payload, cls=plotly.utils.PlotlyJSONEncoder)
    return Response(body, mimetype='application/json')

@app.route('/train_knn')
def train_knn():
    return render_template('train_knn.html')

# KNN Training API
@app.route('/api/train_knn', methods=['POST'])
def api_train_knn():
    features = session.get('selected_features', [])
    target = session.get('selected_target', None)
    cv_folds = int(session.get('cv_folds', 5))
    train_filename = session.get('train_file')
    train_path = os.path.join(UPLOAD_FOLDER, train_filename)

    df = pd.read_csv(train_path)
    X = df[features]
    y = df[target]

    knn = KNeighborsClassifier(n_neighbors=5)
    cv = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=42)

    accuracy_scores = cross_val_score(knn, X, y, cv=cv, scoring='accuracy')
    precision_scores = cross_val_score(knn, X, y, cv=cv,
                                       scoring=make_scorer(precision_score, average='macro', zero_division=0))
    recall_scores = cross_val_score(knn, X, y, cv=cv,
                                    scoring=make_scorer(recall_score, average='macro', zero_division=0))
    f1_scores = cross_val_score(knn, X, y, cv=cv,
                                 scoring=make_scorer(f1_score, average='macro', zero_division=0))

    result = {
        "accuracy": f"{np.mean(accuracy_scores):.2f}",
        "accuracy_std": f"{np.std(accuracy_scores):.2f}",
        "precision": f"{np.mean(precision_scores):.2f}",
        "recall": f"{np.mean(recall_scores):.2f}",
        "f1": f"{np.mean(f1_scores):.2f}"
    }
    return jsonify(result)
@app.route('/download_results_knn', methods=['POST'])
def download_results_knn():
    data = request.get_json(silent=True) or {}
    metrics = data.get('metrics', {})

    # ---- Gather from session
    model_name   = 'KNN'  # force name for this sheet
    dataset_name = session.get('dataset_name') or session.get('cleaned_file') or session.get('train_file', '—')
    features     = session.get('selected_features', [])
    target       = session.get('selected_target', '—')
    cv_folds     = session.get('cv_folds', 5)
    sr           = session.get('split_ratio') or {}
    train_size   = session.get('train_size', '—')
    test_size    = session.get('test_size', '—')
    preproc_list = session.get('preprocess_summary', [])
    method_code  = session.get('balance_method')
    before       = session.get('class_counts_before', {})
    after        = session.get('class_counts_after', {})

    method_map = {'under': 'Undersampling', 'over': 'Oversampling', 'both': 'Both over and under sampling'}
    balance_method = method_map.get(method_code, 'None' if not method_code else method_code)

    if sr and 'train' in sr and 'test' in sr:
        split_ratio_str = f"{int(round(sr['train']*100))} : {int(round(sr['test']*100))}"
    else:
        split_ratio_str = '—'

    rows = []
    add = rows.append

    # Model Summary
    add(('Model Summary', 'Model', model_name))
    add(('Model Summary', 'Dataset', dataset_name))
    add(('Model Summary', 'Target', target))
    add(('Model Summary', 'Features', ', '.join(features)))
    add(('Model Summary', 'Cross-Validation Folds', cv_folds))

    # Data Split
    add(('Data Split', 'Train/Test Split', split_ratio_str))
    add(('Data Split', 'Train set size', train_size))
    add(('Data Split', 'Test set size', test_size))

    # Class Balancing
    add(('Class Balancing', 'Method', balance_method))
    if before:
        for k, v in sorted(before.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'Before - Class {k}', v))
    if after:
        for k, v in sorted(after.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'After - Class {k}', v))

    # Preprocessing
    if preproc_list:
        for i, step in enumerate(preproc_list, 1):
            add(('Preprocessing', f'Step {i}', step))
    else:
        add(('Preprocessing', 'Steps', 'None'))

    # Results
    add(('Results', 'Accuracy',      metrics.get('accuracy')))
    add(('Results', 'Accuracy Std',  metrics.get('accuracy_std')))
    add(('Results', 'Precision',     metrics.get('precision')))
    add(('Results', 'Recall',        metrics.get('recall')))
    add(('Results', 'F1-score',      metrics.get('f1')))

    df_report  = pd.DataFrame(rows, columns=['Section', 'Field', 'Value'])
    df_summary = df_report[df_report['Section'] == 'Model Summary'][['Field', 'Value']]
    df_metrics = df_report[df_report['Section'] == 'Results'][['Field', 'Value']].T
    df_metrics.columns = df_metrics.iloc[0]; df_metrics = df_metrics[1:].reset_index(drop=True)
    df_preproc = pd.DataFrame({'Preprocessing Steps': preproc_list or ['None']})
    classes = sorted(set(list(before.keys()) + list(after.keys())))
    df_balance = pd.DataFrame({
        'Class': classes,
        'Before': [before.get(c, 0) for c in classes],
        'After':  [after.get(c, 0) for c in classes],
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_report.to_excel(writer, index=False, sheet_name='Report (All-in-one)')
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        df_metrics.to_excel(writer, index=False, sheet_name='Metrics')
        df_preproc.to_excel(writer, index=False, sheet_name='Preprocessing')
        df_balance.to_excel(writer, index=False, sheet_name='Class Balance')
        pd.DataFrame({'Features': features}).to_excel(writer, index=False, sheet_name='Features')

        ws = writer.sheets['Report (All-in-one)']
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 80

        from openpyxl.styles import Alignment
        left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = left

        for name in ['Summary', 'Metrics', 'Preprocessing', 'Class Balance', 'Features']:
            if name in writer.sheets:
                s = writer.sheets[name]
                for row in s.iter_rows(min_row=1, max_row=s.max_row, min_col=1, max_col=s.max_column):
                    for cell in row:
                        cell.alignment = left

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='knn_results.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Plot: Error vs K

@app.route('/plot/error_vs_k')
def plot_error_vs_k():
    train_filename = session['train_file']
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_filename))
    X = df[session['selected_features']]
    y = df[session['selected_target']]

    X_train, X_val, y_train, y_val = train_test_split(
        X, y, test_size=0.4, random_state=42, stratify=y
    )

    ks = list(range(1, 31))
    train_err = []
    val_err = []
    for k in ks:
        model = KNeighborsClassifier(n_neighbors=k)
        model.fit(X_train, y_train)
        train_err.append(1 - model.score(X_train, y_train))
        val_err.append(1 - model.score(X_val, y_val))

    # Build Plotly figure
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=ks, y=train_err, mode='lines+markers', name='Train Error'))
    fig.add_trace(go.Scatter(x=ks, y=val_err, mode='lines+markers', name='Validation Error'))
    fig.update_layout(
        title='Error vs. k',
        xaxis_title='Number of Neighbors (k)',
        yaxis_title='Error Rate',
        margin=dict(l=40, r=40, t=50, b=40)
    )

    return Response(
        fig.to_json(),
        mimetype='application/json'
    )
@app.route('/plot/decision_boundary')
def plot_decision_boundary():
    train_filename = session['train_file']
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_filename))
    X = df[session['selected_features']]
    y = df[session['selected_target']]

    # --- Downsample for visualization only
    if len(X) > 2000:
        sample_df = df.sample(n=2000, random_state=42)
        X = sample_df[session['selected_features']]
        y = sample_df[session['selected_target']]

    # Encode labels ONLY for plotting (keep training labels as-is)
    from sklearn.preprocessing import LabelEncoder
    le = None
    if y.dtype == 'O' or isinstance(y.iloc[0], str):
        le = LabelEncoder()
        y_encoded = le.fit_transform(y)
    else:
        y_encoded = y.to_numpy()

    model = KNeighborsClassifier(n_neighbors=5)
    model.fit(X, y)  # training unchanged

    pca = PCA(n_components=2)
    X2d = pca.fit_transform(X)

    x_min, x_max = X2d[:, 0].min() - 1, X2d[:, 0].max() + 1
    y_min, y_max = X2d[:, 1].min() - 1, X2d[:, 1].max() + 1
    xx, yy = np.meshgrid(
        np.linspace(x_min, x_max, 200),
        np.linspace(y_min, y_max, 200)
    )
    grid = np.c_[xx.ravel(), yy.ravel()]

    # Predict with feature names to avoid the warning
    grid_orig = pca.inverse_transform(grid)
    grid_df = pd.DataFrame(grid_orig, columns=X.columns)
    pred = model.predict(grid_df)

    # Convert predicted class labels to ints for the contour
    if le is not None:
        Z = le.transform(pred).reshape(xx.shape)
    else:
        Z = pred.reshape(xx.shape).astype(float)

        # --- prepare list-typed axes that exactly match z's shape ---
    x_vals = xx[0, :].tolist()                  # length == Z.shape[1]
    y_vals = yy[:, 0].tolist()                  # length == Z.shape[0]
    Z_plot = Z.astype(float).tolist()           # 2D list

    fig = go.Figure([
        go.Contour(
            x=x_vals,
            y=y_vals,
            z=Z_plot,
            showscale=False,
            opacity=0.35,
            contours=dict(coloring='heatmap')
        ),
        go.Scatter(
            x=X2d[:, 0].tolist(),
            y=X2d[:, 1].tolist(),
            mode='markers',
            marker=dict(
                size=6,
                color=(y_encoded.tolist() if hasattr(y_encoded, "tolist") else y_encoded),
                showscale=False
            ),
            name='Points'
        )
    ])

    fig.update_layout(
        title='KNN Decision Boundary (2D PCA)',
        xaxis_title='PC1',
        yaxis_title='PC2',
        margin=dict(l=40, r=40, t=50, b=40)
    )

    return Response(fig.to_json(), mimetype='application/json')


# Plot: Confusion Matrix

@app.route('/plot/knn_confusion_matrix')
def plot_knn_confusion_matrix():
    train_filename = session['train_file']
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_filename))
    X = df[session['selected_features']]
    y = df[session['selected_target']]

    preds = cross_val_predict(
        KNeighborsClassifier(n_neighbors=5),
        X, y,
        cv=int(session['cv_folds'])
    )
    cm = confusion_matrix(y, preds)
    labels = [str(c) for c in np.unique(y)]

    fig = go.Figure(go.Heatmap(
        z=cm.tolist(),
        x=labels,
        y=labels,
        colorscale='Blues',
        hoverongaps=False,
        showscale=True
    ))
    fig.update_layout(
        title='KNN Confusion Matrix',
        xaxis=dict(title='Predicted', type='category'),
        yaxis=dict(title='Actual', type='category', autorange='reversed'),
        margin=dict(l=60, r=60, t=50, b=50)
    )

    return Response(
        fig.to_json(),
        mimetype='application/json'
    )

# Render the new template
@app.route('/train_logistic')
def train_logistic():
    return render_template('train_logistic.html')

# API: train logistic regression
@app.route('/api/train_logistic', methods=['POST'])
def api_train_logistic():
    features = session.get('selected_features', [])
    target   = session.get('selected_target', None)
    cv_folds = int(session.get('cv_folds', 5))
    train_filename = session.get('train_file')
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_filename))
    X, y = df[features], df[target]

    clf = LogisticRegression(max_iter=1000, solver='lbfgs')
    cv = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=42)

    accs = cross_val_score(clf, X, y, cv=cv, scoring='accuracy')
    precs= cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(precision_score, average='macro', zero_division=0))
    recs = cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(recall_score,    average='macro', zero_division=0))
    f1s  = cross_val_score(clf, X, y, cv=cv, scoring=make_scorer(f1_score,         average='macro', zero_division=0))
    # Fit final model on all data for plotting
    # --- RULE B1: Get out-of-fold probabilities via CV to match your CV metrics ---
    # These are predictions for each row made by a model that did NOT see that row.
    from sklearn.model_selection import cross_val_predict
    probs_cv = cross_val_predict(clf, X, y, cv=cv, method='predict_proba')[:, 1]

    # --- RULE B2: Save labels and probabilities to session for plotting the ROC later ---
    # Keep the labels as-is; we'll encode them in the plotting route to be safe.
    session['logistic_y_true'] = np.asarray(y).tolist()
    session['logistic_proba']  = np.asarray(probs_cv).tolist()
    # -------------------------------------------------------------------------------

    clf.fit(X, y)
    session['logistic_model'] = pickle.dumps(clf)  # save model to session
    return jsonify({
      "accuracy":      f"{np.mean(accs):.2f}",
      "accuracy_std":  f"{np.std(accs):.2f}",
      "precision":     f"{np.mean(precs):.2f}",
      "recall":        f"{np.mean(recs):.2f}",
      "f1":            f"{np.mean(f1s):.2f}"
    })

@app.route('/download_results_logistic', methods=['POST'])
def download_results_logistic():
    from openpyxl.styles import Alignment  # already imported at top, safe to keep here too
    data = request.get_json(silent=True) or {}
    metrics = data.get('metrics', {})

    # Build from session
    model_name   = 'Logistic Regression'
    dataset_name = session.get('dataset_name') or session.get('cleaned_file') or session.get('train_file', '—')
    features     = session.get('selected_features', [])
    target       = session.get('selected_target', '—')
    cv_folds     = session.get('cv_folds', 5)
    sr           = session.get('split_ratio') or {}
    train_size   = session.get('train_size', '—')
    test_size    = session.get('test_size', '—')
    preproc_list = session.get('preprocess_summary', [])
    method_code  = session.get('balance_method')
    before       = session.get('class_counts_before', {})
    after        = session.get('class_counts_after', {})

    # Friendly class-balance method
    method_map = {'under': 'Undersampling', 'over': 'Oversampling', 'both': 'Both over and under sampling'}
    balance_method = method_map.get(method_code, 'None' if not method_code else method_code)

    # Split ratio as 80 : 20
    if sr and 'train' in sr and 'test' in sr:
        split_ratio_str = f"{int(round(sr['train']*100))} : {int(round(sr['test']*100))}"
    else:
        split_ratio_str = '—'

    # Compose rows
    rows = []
    add = rows.append
    add(('Model Summary', 'Model', model_name))
    add(('Model Summary', 'Dataset', dataset_name))
    add(('Model Summary', 'Target', target))
    add(('Model Summary', 'Features', ', '.join(features)))
    add(('Model Summary', 'Cross-Validation Folds', cv_folds))

    add(('Data Split', 'Train/Test Split', split_ratio_str))
    add(('Data Split', 'Train set size', train_size))
    add(('Data Split', 'Test set size', test_size))

    add(('Class Balancing', 'Method', balance_method))
    if before:
        for k, v in sorted(before.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'Before - Class {k}', v))
    if after:
        for k, v in sorted(after.items(), key=lambda kv: str(kv[0])):
            add(('Class Balancing', f'After - Class {k}', v))

    if preproc_list:
        for i, step in enumerate(preproc_list, 1):
            add(('Preprocessing', f'Step {i}', step))
    else:
        add(('Preprocessing', 'Steps', 'None'))

    add(('Results', 'Accuracy',     metrics.get('accuracy')))
    add(('Results', 'Accuracy Std', metrics.get('accuracy_std')))
    add(('Results', 'Precision',    metrics.get('precision')))
    add(('Results', 'Recall',       metrics.get('recall')))
    add(('Results', 'F1-score',     metrics.get('f1')))

    # DataFrames
    df_report  = pd.DataFrame(rows, columns=['Section', 'Field', 'Value'])
    df_summary = df_report[df_report['Section']=='Model Summary'][['Field','Value']]
    df_metrics = df_report[df_report['Section']=='Results'][['Field','Value']].T
    if not df_metrics.empty:
        df_metrics.columns = df_metrics.iloc[0]
        df_metrics = df_metrics[1:].reset_index(drop=True)
    df_preproc = pd.DataFrame({'Preprocessing Steps': preproc_list or ['None']})
    classes = sorted(set(list(before.keys()) + list(after.keys())))
    df_balance = pd.DataFrame({
        'Class': classes,
        'Before': [before.get(c, 0) for c in classes],
        'After':  [after.get(c, 0) for c in classes],
    })
    df_features = pd.DataFrame({'Features': features})

    # Write Excel
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_report.to_excel(writer, index=False, sheet_name='Report (All-in-one)')
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        df_metrics.to_excel(writer, index=False, sheet_name='Metrics')
        df_preproc.to_excel(writer, index=False, sheet_name='Preprocessing')
        df_balance.to_excel(writer, index=False, sheet_name='Class Balance')
        df_features.to_excel(writer, index=False, sheet_name='Features')

        ws = writer.sheets['Report (All-in-one)']
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 80

        left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = left
        for name in ['Summary','Metrics','Preprocessing','Class Balance','Features']:
            if name in writer.sheets:
                s = writer.sheets[name]
                for row in s.iter_rows(min_row=1, max_row=s.max_row, min_col=1, max_col=s.max_column):
                    for cell in row:
                        cell.alignment = left

    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='logistic_regression_results.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Plot 1: ROC Curve
@app.route('/plot/roc_curve')
def plot_roc_curve():
    try:
        import pickle
        train_filename = session.get('train_file')
        if not train_filename:
            return jsonify(error="Train file not set in session"), 400

        file_path = os.path.join(UPLOAD_FOLDER, train_filename)
        if not os.path.exists(file_path):
            return jsonify(error=f"File not found: {train_filename}"), 400

        df = pd.read_csv(file_path)
        X = df[session['selected_features']]
        y = df[session['selected_target']]

        # ROC needs binary target
        classes = np.unique(y)
        if len(classes) != 2:
            return jsonify(error=f"ROC curve requires binary classification but found {len(classes)} classes"), 400

        # Encode labels if object
        if y.dtype == 'O':
            from sklearn.preprocessing import LabelEncoder
            y = LabelEncoder().fit_transform(y)
        
        cv = StratifiedKFold(n_splits=int(session.get('cv_folds', 5)), shuffle=True, random_state=42)

        # Prefer the CV probabilities saved during training (matches your CV metrics)
        if 'logistic_proba' in session:
            probs = np.array(session['logistic_proba'], dtype=float)
        else:
          # Fallback to your current behavior
            if 'logistic_model' in session:
                clf = pickle.loads(session['logistic_model'])
                probs = cross_val_predict(clf, X, y, cv=cv, method='predict_proba')[:, 1]
            else:
                clf = LogisticRegression(max_iter=5000, solver='lbfgs')
                probs = cross_val_predict(clf, X, y, cv=cv, method='predict_proba')[:, 1]


        # Compute ROC
        fpr, tpr, _ = roc_curve(y, probs)
        roc_auc = auc(fpr, tpr)

        fig = go.Figure()
        fig.add_shape(
            type='line', x0=0, y0=0, x1=1, y1=1,
            line=dict(dash='dash', color='gray'),
            layer='below'
        )
        fig.add_trace(go.Scatter(
            x=np.asarray(fpr, dtype=float).tolist(),
            y=np.asarray(tpr, dtype=float).tolist(),
            mode='lines',
            name=f'ROC (AUC = {roc_auc:.2f})'
        ))


        subtitle = " (AUC ≈ 0.5: model ~ random)" if abs(roc_auc - 0.5) < 1e-3 else ""
        fig.update_layout(
            title=f'ROC Curve{subtitle}',
            xaxis_title='False Positive Rate',
            yaxis_title='True Positive Rate',
            margin=dict(l=40, r=40, t=50, b=40)
        )

        fig_json = fig.to_plotly_json()
        return Response(json.dumps({"data": fig_json["data"], "layout": fig_json["layout"]},
                                   cls=plotly.utils.PlotlyJSONEncoder),
                        mimetype='application/json')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(error=str(e)), 500

# Plot 2: Precision–Recall Curve
@app.route('/plot/precision_recall')
def plot_precision_recall():
    try:
        train_filename = session.get('train_file')
        if not train_filename:
            return jsonify(error="Train file not set in session"), 400

        file_path = os.path.join(UPLOAD_FOLDER, train_filename)
        if not os.path.exists(file_path):
            return jsonify(error=f"File not found: {train_filename}"), 400

        df = pd.read_csv(file_path)
        X = df[session['selected_features']]
        y = df[session['selected_target']]

        # ✅ Normalize binary labels to 0/1
        if y.dtype != 'O' and np.issubdtype(y.dtype, np.number) and len(np.unique(y)) == 2:
            unique_vals = np.unique(y)
            if not np.array_equal(unique_vals, [0, 1]):
                from sklearn.preprocessing import LabelEncoder
                y = LabelEncoder().fit_transform(y)
            else:
                y = np.asarray(y, dtype=int)

        # PR (this version) expects binary labels
        classes = np.unique(y)
        if len(classes) != 2:
            return jsonify(error=f"Precision–Recall curve requires binary classification but found {len(classes)} classes"), 400

        if y.dtype == 'O':
            from sklearn.preprocessing import LabelEncoder
            y = LabelEncoder().fit_transform(y)

        clf = LogisticRegression(max_iter=1000, solver='lbfgs')
        cv = StratifiedKFold(n_splits=int(session.get('cv_folds', 5)), shuffle=True, random_state=42)

        # CV probabilities for the positive class
        probs = cross_val_predict(clf, X, y, cv=cv, method='predict_proba')[:, 1]

        precision, recall, _ = precision_recall_curve(y, probs)
        pr_auc = auc(recall, precision)

        # NEW: convert to lists so Plotly always renders
        prec_list = np.asarray(precision, dtype=float).tolist()
        rec_list  = np.asarray(recall,    dtype=float).tolist()

        # --- build ONE figure (don't overwrite it) ---
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=rec_list,
            y=prec_list,
            mode='lines',
            name=f'AUC = {pr_auc:.2f}'
        ))
        fig.update_layout(
            title='Precision–Recall Curve',
            xaxis_title='Recall',
            yaxis_title='Precision',
            margin=dict(l=40, r=40, t=50, b=40)
        )

        fig_json = fig.to_plotly_json()
        body = json.dumps({"data": fig_json["data"], "layout": fig_json["layout"]},
                          cls=plotly.utils.PlotlyJSONEncoder)
        return Response(body, mimetype='application/json')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(error=str(e)), 500


# Plot 3: logistic_confusion_matrix
@app.route('/plot/logistic_confusion_matrix')
def plot_logistic_confusion_matrix():
    try:
        train_filename = session.get('train_file')
        if not train_filename:
            return jsonify(error="Train file not set in session"), 400

        train_path = os.path.join(UPLOAD_FOLDER, train_filename)
        if not os.path.exists(train_path):
            return jsonify(error=f"File not found: {train_filename}"), 400

        df = pd.read_csv(train_path)
        X = df[session['selected_features']]
        y = df[session['selected_target']]

        # Ensure stratified CV and valid number of splits
        counts = pd.Series(y).value_counts()
        if len(counts) < 2:
            return jsonify(error="Confusion matrix needs at least two classes in the target"), 400

        requested = int(session.get('cv_folds', 5))
        min_class = int(counts.min())
        n_splits = max(2, min(requested, min_class))  # cap by minority class

        cv = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)

        clf = LogisticRegression(max_iter=1000, solver='lbfgs')
        preds = cross_val_predict(clf, X, y, cv=cv)

        # Keep label order consistent
        labels_unique = np.unique(y)
        cm = confusion_matrix(y, preds, labels=labels_unique)
        labels = [str(c) for c in labels_unique]

        fig = go.Figure(go.Heatmap(
            z=cm.tolist(),
            x=labels,
            y=labels,
            colorscale='Blues',
            hoverongaps=False,
            showscale=True
        ))
        fig.update_layout(
            title=f'Logistic Regression Confusion Matrix (CV={n_splits})',
            xaxis=dict(title='Predicted', type='category'),
            yaxis=dict(title='Actual', type='category', autorange='reversed'),
            margin=dict(l=60, r=60, t=50, b=50)
        )

        js = fig.to_plotly_json()
        body = json.dumps({"data": js["data"], "layout": js["layout"]},
                          cls=plotly.utils.PlotlyJSONEncoder)
        return Response(body, mimetype='application/json')

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(error=str(e)), 500
# ---- Generic Experiment endpoints (one GET + one POST) ----

@app.route("/experiment/<model_name>", methods=["GET"])
def show_experiment_generic(model_name):
    if not session.get('selected_features'):
        return "Please select features and train first.", 400
    if not uploaded_file_path or not os.path.exists(uploaded_file_path):
        return "Please upload data first.", 400

    # original file
    df_raw = pd.read_csv(uploaded_file_path)

    # mirror "Clean column names" if used
    steps = (session.get('preprocess') or {}).get('steps', [])
    if 'clean' in steps:
        df_raw.columns = _normalize_names(df_raw.columns)

    # map encoded feature names back to their source columns (for one-hot)
    encoded_to_source = session.get('encoded_to_source', {}) or {}
    selected_clean = session.get('selected_features', [])
    source_cols = []
    for col in selected_clean:
        src = encoded_to_source.get(col, col)
        if src not in source_cols:
            source_cols.append(src)

    # only keep columns present in original df
    source_cols = [c for c in source_cols if c in df_raw.columns]

    schema = _build_schema_from_df(df_raw, source_cols)
    return render_template("experiment_generic.html", schema=schema, model_name=model_name)
@app.route("/experiment/<model_name>/predict", methods=["POST"])
def experiment_generic_predict(model_name):
    payload = request.get_json(force=True)

    train_file = session['train_file']
    df = pd.read_csv(os.path.join(UPLOAD_FOLDER, train_file))
    X = df[session['selected_features']]
    y = df[session['selected_target']]

    if model_name == "decision_tree":
        model = DecisionTreeClassifier(random_state=42)
    elif model_name == "knn":
        model = KNeighborsClassifier(n_neighbors=5)
    elif model_name in ("logistic", "logistic_regression"):
        model = LogisticRegression(max_iter=1000, solver='lbfgs')
    else:
        return jsonify(error=f"Unknown model: {model_name}"), 400

    model.fit(X, y)

    row = _payload_to_model_row(payload)

    pred = model.predict(row)[0]
    try:
        pred_val = int(pred)
    except Exception:
        pred_val = str(pred)

    if hasattr(model, "predict_proba"):
        probs = model.predict_proba(row)[0].tolist()
        return jsonify(prediction=pred_val, probabilities=probs)

    return jsonify(prediction=pred_val)


if __name__ == '__main__':
    app.run(debug=True)
